# -*- coding: utf-8 -*-
"""
VCIA
for python 3
@author: igor bertyaev

программа берет на вход файлы not gabarit насчитанные террасканом
и формирует выходную таблицу установленного образца
программа также удаляет дублирующиеся координаты нерушений, 
оставляя одно значение с наименьшим габаритом

"""

import os
import fnmatch
import openpyxl
from math import sqrt


fz = []                # list of not_gabarit files
ctow = 'нет !'
spec = 'нет !'
logo = 'нет !'

##### looking for files ######

for file in os.listdir('.'):
    if fnmatch.fnmatch(file, '*not_gabarit*.txt'):    #fulling list of not gabarit files
        fz.append(file)
    if fnmatch.fnmatch(file, '*tow*.pts'):    #finding cgtow file
        ctow = file              
    if fnmatch.fnmatch(file, '*pecifica*.xlsx'):    #finding cgtow file
        spec = file              
    if fnmatch.fnmatch(file, '*logo*'):    #finding cgtow file
        logo = file              


#  вступительная речь, интервью, интерфейс
print('\nпривет! поработаем?\n ')
print('найдены файлы негабаритов на вход:') 
print('\n'.join(fz))
print('координаты опор: ',ctow)
print('Specification: ',spec)
print('логотип для таблицы: ',logo)
print('\nдалее будет произведен расчет, ')
print('чтобы продолжить нажмите 1, чтобы выйти нажмите 2\n')

vr = 0           
while vr not in range(1,3):
    try:
        vr = int(input('введите 1 или 2 и нажмите Enter: '))
    except ValueError:
        print('не похоже на цифры')
    if vr not in range(1,5):
        print('надо выбрать 1 или 2!')
            
if vr == 1:
    if len(fz) < 1:
        input('\nне найдены файлы *not_gabarit*, программа будет закрыта')
        quit()
    if ctow == 'нет !':
        input('\nне найден файл *ctow*, программа будет закрыта')
        quit()
    if spec == 'нет !':
        input('\nне найден файл *Specification*, программа будет закрыта')
        quit()
    if logo == 'нет !':
        print('\nне найден логотип, выходная таблица будет рассчитана без него')
        input('нажмите Enter для продолжения')

elif vr == 2:
    quit()
else:
    print('number error')

################  looking for name of line 

name_split = spec.split("_")
num_name = str(name_split[0] + '_' + name_split[1])
#print(num_name)


##################

corr_list = []      # list without duplicates
global len_ini
len_ini = 0         # number of stroke in initial files

if not os.path.exists('out_vcia'):
    os.makedirs('out_vcia')                # make a directory 'out_pts'

for fz_n in range(len(fz)):
        Fzx = fz[fz_n]
        infile = open(Fzx, 'r')            # choose file and open it
        firstline = infile.readline().strip()         # первую строку пишем отдельно чтобы лист был не пустой 
        corr_list.append(firstline.split("\t"))             
        len_ini += 1
        
        for line in infile:
            line = line.rstrip()             # удаляем символы конца строки
            words = line.split("\t")
            len_ini += 1
           
            if words[0] == corr_list[-1][0]:
                tr = 0
                for pt in range(len(corr_list)):

                    if words[3] == corr_list[pt][3] and words[4] == corr_list[pt][4] and words[5] == corr_list[pt][5]:
                        ## проверка меньшего числа на случай другой сортировки
                        if words[2] > corr_list[pt][2]:
                            None
                        else:
                            corr_list[pt] = words  ## !! проверить как работает
                    else:
                        tr += 1
                        if tr == len(corr_list):
                            corr_list.append(words)
                        else:
                            None
            else:                
                corr_list.append(words)
        infile.close


#print('удалено дублирующихся строк: '+ str(len_ini - len(corr_list)))        
        
#############
           
outfile = open(str('out_vcia/not_gabarit_corr.txt'), 'w')

for a in range(len(corr_list)):
    for b in range(6):
        outfile.write(corr_list[a][b]+'\t')
    outfile.write('\n')

outfile.close()            


'''
############  list of used towers.  not used
def used_tows_f():
    used_tows = []
    for t in range(len(corr_list)):
        if corr_list[t][0] not in used_tows:
            used_tows.append(corr_list[t][0])
    
    used_tows = list(map(int, used_tows))
    
    for n in range(len(used_tows)):
        if used_tows[n]-1 != used_tows[n-1]:
            used_tows.append(int(used_tows[n-1]+1))
    
    used_tows.sort()

#print(used_tows)
'''

###########  add coordinates of towers

l_ctow = []    # list of ctow from file

infile = open(ctow, 'r')            # choose file and open it
for line in infile:
    line = line.rstrip()             # удаляем символы конца строки
    words = line.split()
    l_ctow.append(words)

infile.close

#########################################  add span lenght

for l in range(len(l_ctow)-1):
    if int(l_ctow[l+1][0]) == int(l_ctow[l][0])+1:
        x1 = float(l_ctow[l][1])
        y1 = float(l_ctow[l][2])
    #    z1 = float(tow_list[l][3])    
        x2 = float(l_ctow[l+1][1])    
        y2 = float(l_ctow[l+1][2])    
    #    z2 = float(l_ctow[l+1][3])    
    #    dist3d = sqrt((x1-x2)**2+(y1-y2)**2+(z1-z2)**2)
        dist = sqrt((x1-x2)**2+(y1-y2)**2)
        l_ctow[l].append(round(dist, 2))
    else:
        l_ctow[l].append('')
l_ctow[len(l_ctow)-1].append('')    

#print(l_ctow)

a = 0
for a in range(len(corr_list)):
    corr_list[a][1] = str(int(corr_list[a][0])+1)
    for coo in range(len(l_ctow)):
        if str(int(corr_list[a][0])+1) == l_ctow[coo][0]:
            corr_list[a].append(l_ctow[coo][1])
            corr_list[a].append(l_ctow[coo][2])
            corr_list[a].append(l_ctow[coo][4])
            corr_list[a].append(l_ctow[coo+1][1])
            corr_list[a].append(l_ctow[coo+1][2])
            
        else:
            None

#print(corr_list)

###########  dict id names

id_tows = {}     # creating dict  *work name = id name*

wb = openpyxl.load_workbook(spec)
sheet = wb.active

for cn in range(3, sheet.max_row):      ## cn - cell number
    if not sheet['A'+(str(cn))].value == sheet['A'+(str(cn+1))].value:
        if isinstance(sheet['A'+(str(cn))].value, int):
            id_tows[str(sheet['A'+(str(cn))].value)] = str(sheet['B'+(str(cn))].value)
    else:
        break
    
#print(id_tows)

#########################  change to id names
a = 0
for a in range(len(corr_list)):
    corr_list[a][0] = id_tows[str(int(corr_list[a][0])+1)]
    corr_list[a][1] = id_tows[str(int(corr_list[a][1])+1)]
    x1 = float(corr_list[a][6])
    y1 = float(corr_list[a][7])
    x2 = float(corr_list[a][9])    
    y2 = float(corr_list[a][10])    
    x0 = float(corr_list[a][3])    
    y0 = float(corr_list[a][4])    
    dist_A = float(corr_list[a][8])
    dist_B = ((y2-y1)*x0-(x2-x1)*y0+x2*y1-y2*x1)/dist_A
    dist_C = sqrt((x0-x1)**2+(y0-y1)**2)
    dist_D = sqrt(dist_C**2-dist_B**2)
    
    corr_list[a].append(round(dist_B, 2))
    corr_list[a].append(round(dist_D, 2))
    corr_list[a].append(str(name_split[0] + '_'+ corr_list[a][0]+ '-'+ corr_list[a][1] + '_01.jpg'))

           
outfile = open(str('out_vcia/not_gabarit_corr_v2.txt'), 'w')

a = 0
b = 0
for a in range(len(corr_list)):
    for b in range(len(corr_list[a])):
        outfile.write(str(corr_list[a][b])+'\t')
    outfile.write('\n')

outfile.close()      
#print(corr_list)


################  work with excel

wb = openpyxl.Workbook()
sheet = wb.active

sheet.title = num_name

### looking for logo and place it 
if logo == 'нет !':
    None
else:
    img = openpyxl.drawing.image.Image(logo)  
    sheet.add_image(img, 'F2')

### fill the file 
sheet['A2'] = 'Line name:'
sheet['B2'] = num_name
sheet['A3'] = 'Date of survey:'

sheet.freeze_panes = 'C7'
sheet.sheet_view.zoomScale = 90

sheet.merge_cells('A5:A6')
sheet['A5'] = 'From tower'
sheet.merge_cells('B5:B6')
sheet['B5'] = 'To tower'
sheet.merge_cells('C5:C6')
sheet['C5'] = 'Span length, m'
sheet.merge_cells('D5:D6')
sheet['D5'] = 'Conductor Temperature, ºС'
sheet.merge_cells('E5:E6')
sheet['E5'] = 'Clearance to infringing tree, m'
sheet.merge_cells('F5:G5')
sheet['F5'] = 'Infringing tree location'
sheet['F6'] = 'Station, m'
sheet['G6'] = 'Offset, m'
sheet.merge_cells('H5:H6')
sheet['H5'] = 'Photography Number'
     
   
     
for ce in range(len(corr_list)):
    sheet[str('A'+str(int(7+ce)))] = corr_list[ce][0]
    sheet[str('B'+str(int(7+ce)))] = corr_list[ce][1]
    sheet[str('C'+str(int(7+ce)))] = corr_list[ce][8]
    sheet[str('E'+str(int(7+ce)))] = float(corr_list[ce][2])
    sheet[str('F'+str(int(7+ce)))] = corr_list[ce][12]
    sheet[str('G'+str(int(7+ce)))] = corr_list[ce][11]
    sheet[str('H'+str(int(7+ce)))] = corr_list[ce][13]
    
### a little bit of style
from openpyxl.styles import Border, Side, Alignment, Font

font1 = Font(name='Arial', size=12)
font1b = Font(name='Arial', size=12, bold=True)
font2 = Font(name='Arial', size=10, bold=True)
font3 = Font(name='Arial', size=10)

align1 = Alignment(horizontal='center', vertical='center', wrap_text = True)

bold_left = Border(left=Side(style='medium'), 
                     right=Side(style='thin'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

bold_right = Border(left=Side(style='thin'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

bold_mid = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

bold_top_mid = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='thin'))

bold_bot_mid = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='medium'))

bold_mid_left = Border(left=Side(style='medium'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

bold_mid_right = Border(left=Side(style='thin'), 
                     right=Side(style='medium'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

bold_left_bot = Border(left=Side(style='medium'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='medium'))

bold_right_bot = Border(left=Side(style='thin'), 
                     right=Side(style='medium'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='medium'))


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 10
sheet.column_dimensions['D'].width = 13
sheet.column_dimensions['E'].width = 13                       
sheet.column_dimensions['F'].width = 13                       
sheet.column_dimensions['G'].width = 13                       
sheet.column_dimensions['H'].width = 35                       
sheet.row_dimensions[5].height = 30                       
#sheet.row_dimensions[6].height = 15                       
                
                        
sheet['A2'].font = font1
sheet['B2'].font = font1b
sheet['A3'].font = font1

sheet['A5'].border = bold_left
sheet['A6'].border = bold_left
sheet['B5'].border = bold_mid
sheet['B6'].border = bold_mid
sheet['C5'].border = bold_mid
sheet['C6'].border = bold_mid
sheet['D5'].border = bold_mid
sheet['D6'].border = bold_mid
sheet['E5'].border = bold_mid
sheet['E6'].border = bold_mid
sheet['F5'].border = bold_top_mid
sheet['F6'].border = bold_bot_mid
sheet['G5'].border = bold_top_mid
sheet['G6'].border = bold_bot_mid
sheet['H5'].border = bold_right
sheet['H6'].border = bold_right
     
     
for row in sheet['A5':'H6']:
    for cell in row:
        cell.alignment = align1
        cell.font = font2
        
        
for row in sheet['A7':str('H'+str(len(corr_list)+6))]:
    for cell in row:
        cell.alignment = align1
        cell.font = font3
        cell.border = thin_border

for col in sheet['C7':str('C'+str(len(corr_list)+6))]:
    for cell in col:
        cell.number_format = '0.00'

for col in sheet['E7':str('G'+str(len(corr_list)+6))]:
    for cell in col:
        cell.number_format = '0.00'

for col in sheet['A7':str('A'+str(len(corr_list)+5))]:
    for cell in col:
        cell.border = bold_mid_left

for col in sheet['H7':str('H'+str(len(corr_list)+5))]:
    for cell in col:
        cell.border = bold_mid_right

for row in sheet[str('B'+str(len(corr_list)+6)):str('G'+str(len(corr_list)+6))]:
    for cell in row:
        cell.border = bold_bot_mid
        
sheet[str('A'+str(len(corr_list)+6))].border = bold_left_bot        
sheet[str('H'+str(len(corr_list)+6))].border = bold_right_bot        
  
wb.save(str('out_vcia/'+ num_name +'_VCIA_Report_v01.xlsx'))





print('\nвсё сработало хорошо :) \nфайлы созданы в директории out_vcia\nудалено дублирующихся строк: '+ str(len_ini - len(corr_list))) 
input('нажмите Enter для выхода')

