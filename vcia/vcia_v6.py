# -*- coding: utf-8 -*-
"""
VCIA
for python 3
@author: igor bertyaev

программа для расчета таблицы негабаритов и подготовки данных для QGIS

на вход файлы
- not_gabarit_1.txt - насчитанные террасканом
- stow.pts
- имя_линии_specification.xlsx (XLSX ! - спецификация)

формируется выходная таблица установленного образца
программа также удаляет дублирующиеся координаты нерушений,
оставляя одно значение с наименьшим габаритом

для QGIS формируется набор данных с помощью которых затем строится атлас изображений
qgis1 columns (separated by tab):
(1) work id; (2) id; (3) number of cline; (4, 5, 6) x, y, z

qgis2 (ng_tab3) columns:
(1) ts span name; (2) wire num; (3) 3d dist to ngab; (4,5,6) x,y,z of ngab point;
(7) cline num; (8) work id start span; (9) id span start; (10) id span end;
(11) span length; (12) station; (13) offset; (14) image name

qgis3 columns:
(1) span name id; (2) min ngdist for span; (3) L; (4,5,6) x,y,z of start structure;
(7,8,9) x,y,z of end structure; (10) azimuth; (11) map angle, (12) num of violations

"""

from pathlib import Path
import openpyxl
import csv
from math import sqrt, acos, degrees

p = Path.cwd()  # work dir


def filecollect():
    # collect all the files we need

    ng_fz = sorted(p.glob('*not_gabarit*.txt'))   # filling list of not_gab files

    # get all files using mask and if there are more than 1 writing error
    # if there are no files - name it 'no_file'
    ctow_file = list(p.glob('*tow.pts'))
    if len(ctow_file) == 1:
        ctow_file = ctow_file[0]
    elif len(ctow_file) == 0:
        ctow_file = 'no_file'
    else:
        ctow_file = 'error'

    spec_file = list(p.glob('*pecifica*.xlsx'))
    if len(spec_file) == 1:
        spec_file = spec_file[0]
    elif len(spec_file) == 0:
        spec_file = 'no_file'
    else:
        spec_file = 'error'

    if Path(p / 'logo.png').exists():
        logo_file = p / 'logo.png'
    else:
        logo_file = 'нет !'

    return ng_fz, ctow_file, spec_file, logo_file


def iface():
    #  вступительная речь, интервью, интерфейс
    print('\nпривет! поработаем?\n ')
    print(f'имя линии: {line_name}')
    print('найдены файлы негабаритов на вход:')
    for f in fz:
        print(f'\t- {f.name}')

    if ctow == 'error':
        print('! - ошибка поиска файла *ctow* - найдено больше одного файла')
    elif ctow == 'no_file':
        print('! - ошибка поиска файла *ctow* - файл не найден')
    else:
        print(f'файл координат опор: \n\t- {ctow.name}')

    if spec == 'error':
        print('! - ошибка поиска файла Specification - найдено больше одного файла')
    elif spec == 'no_file':
        print('! - ошибка поиска файла Specification - файл не найден')
        print('! - файл должен быть сохранен как .xlsx !')
    else:
        print(f'Specification: \n\t- {spec.name}')

    print(f'логотип для таблицы: \n\t- logo.png')
    print('\nдалее будет произведен расчет, ')
    print('чтобы продолжить нажмите 1, чтобы выйти нажмите 2\n')

    vr = 0
    while vr not in range(1, 3):
        try:
            vr = int(input('введите 1 или 2 и нажмите Enter: '))
        except ValueError:
            print('не похоже на цифры')
        if vr not in range(1, 5):
            print('надо выбрать 1 или 2!')

    if vr == 1:
        if len(fz) < 1:
            input('\n! - не найдены файлы *not_gabarit*, программа будет закрыта')
            quit()
        if ctow == 'error' or ctow == 'no_file':
            input('\n! - ошибка поиска файла *ctow*, программа будет закрыта')
            quit()
        if spec == 'error' or spec == 'no_file':
            input('\n! - не найден файл *Specification*, программа будет закрыта')
            quit()
        if logo == 'нет !':
            print('\n! - не найден логотип, выходная таблица будет рассчитана без него')
            input('нажмите Enter для продолжения')

    elif vr == 2:
        quit()
    else:
        print('number error')


def centerline(path_str):
    # чтение файла координат опор c рабочей нумерацией и запись его в лист c с удалением дубликатов
    str = []  # list of structures
    with open(path_str) as ffile:
        for line in ffile:
            stroka = line.strip('\n').split()
            # ctow file should contain work-id, x, y, z in each row
            if len(stroka) == 4:
                # check if coords already in list (delete doubles with different names)
                dub = 0
                for i in range(len(str)):
                    # compare x y z written in list
                    if (str[i][1], str[i][2], str[i][3]) == (float(stroka[1]), float(stroka[2]), float(stroka[3])):
                        dub +=1
                if dub < 1:
                    # if there are no such coords, we can write new one
                    str.append([int(stroka[0]), float(stroka[1]), float(stroka[2]), float(stroka[3])])
            elif len(stroka) == 0:
                # if there is blank stroke in the file
                print('blank')
            else:
                # if stroke looks weird
                print('ctow file error')

    print(f'загружено опор: {len(str)}')

    return str


def spec_id(specification):
    # reading xlsx specification and save work and id names of structures in list
    id_tows = []  # creating list * work id, id , num of cline *
    cline_num = 1   # number of cline
    wb = openpyxl.load_workbook(specification)
    sheet = wb.active

    for cn in range(3, sheet.max_row):   ## cn - cell number
        if sheet['A' + (str(cn))].value != sheet['A' + (str(cn + 1))].value:
            if isinstance(sheet['A' + (str(cn))].value, int):
                id_tows.append([int(sheet['A' + (str(cn))].value), str(sheet['B' + (str(cn))].value), cline_num])
            else:
                cline_num += 1   # change cline num after empty stroke in spec
        else:
            break

    return id_tows


def notgabread(list_of_files):
    # read notgab files and create tab without duplicate points and with num of cline
    # exit tab has 7 columns: span num (work ts), wire num, gab, x, y, z, num of cline.

    ngtab = []    # table of notgabs

    for file in list_of_files:
        cline_num = file.stem.split('_')[-1]   # number of cline is the end of filename

        with open(file) as ngfile:
            ngtab.append(ngfile.readline().strip('\n').split())    # first line - add it separately
            ngtab[-1].append(cline_num)    # add cline number to the first stroke
            for line in ngfile:
                stroka = line.strip('\n').split()
                stroka.append(cline_num)   # add cline number
                if stroka[0] == ngtab[-1][0]:
                    # check if same span has same gab point
                    tr = 0   # number of try
                    for pt in range(len(ngtab)):
                        if (stroka[3], stroka[4], stroka[5]) == (ngtab[pt][3], ngtab[pt][4], ngtab[pt][5]):
                            # if all coords the same
                            if stroka[2] < ngtab[pt][2]:
                                # check which one is smaller and keep it
                                ngtab[pt] = stroka
                        else:
                            tr += 1
                            if tr == len(ngtab):
                                # when we read all list and did not find same - take this
                                ngtab.append(stroka)
                else:
                    ngtab.append(stroka)

    return ngtab


def idspannames(ids, notgabtab):
    # in ngtab we should add span names with work and real IDs

    for st in range(len(notgabtab)):
        # first we found start of cline num
        tsn = int(notgabtab[st][0])   # terrascan span num
        cn = int(notgabtab[st][6])   # num of cline
        start_cn = 100000
        for str in range(len(ids)):
            if ids[str][2] == cn:
                if ids[str][0] < start_cn:
                    start_cn = ids[str][0]   # start of this cline

        wid = start_cn + tsn   # now we know work ID
        id_st, id_fin = '-', '-'   # start and finish ID names

        for str2 in range(len(ids)):
            # then found ID names
            if ids[str2][0] == wid:
                id_st = ids[str2][1]
                id_fin= ids[str2 + 1][1]

        # then add work_id, start_id and finish_ID to our table
        for a in [wid, id_st, id_fin]: notgabtab[st].append(a)

    return notgabtab


def calc_h_of_triangle(a, b, c):
    # calc h of triangle
    # if we have triangle with vertices a, b, c
    # then we need to find h - which is start from c and is perpendicular to ab
    # to find it we will use formula:
    # (1) h=2*S/A  where s - area of triangle, A - length of side of triangle (ab in our case)
    # to find S we will use formula: (2) ((xb-xa)*(yc-ya)-(xc-xa)*(yb-ya))/2
    # then use (2) in (1) and have final formula:
    # (3) h = ((xb-xa)*(yc-ya)-(xc-xa)*(yb-ya))/ab
    # ! warning ! in south hemisphere we should use '-' before result,
    # as coordinates goes opposite direction than in north hemisphere

    x1, y1 = a[0], a[1]
    x2, y2 = b[0], b[1]
    x0, y0 = float(c[0]), float(c[1])
    len = round(sqrt((x2-x1)**2+(y2-y1)**2), 2)
    h = round((((x2-x1)*(y0-y1)-(x0-x1)*(y2-y1)) / len), 2)
    off = round(sqrt(((x0-x1)**2+(y0-y1)**2)-h**2), 2)
    return len, -h, off


def filltab(tab, str_coords):
    # now ngtab v2 has 10 columns:
    # (1) ts span name; (2) wire num; (3) 3d dist to ngab; (4,5,6) x,y,z of ngab point; (7) cline num;
    # (8) work id start span; (9) id span start; (10) id span end
    # here we add new columns: (11) span length; (12) station; (13) offset; (14) image name
    for span in range(len(tab)):
        # a = []    #span start coords (x, y)
        # b = []    #span end coords (x, y)
        # c = []    #ngab coords (x, y)
        for struct in range(len(str_coords)):
            if tab[span][7] == str_coords[struct][0]:
                a = str_coords[struct][1:3]
                b = str_coords[struct+1][1:3]
                c = tab[span][3:5]
                l, offs, station = calc_h_of_triangle(a, b, c)
                image_name = f'{tab[span][8]}_{tab[span][9]}.jpg'
                for a in [l, station, offs, image_name]:
                    tab[span].append(a)

    return tab


def excel_export(tab):
    # create excel tab

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.title = line_name

    ### looking for logo and place it
    if logo != 'нет !':
        img = openpyxl.drawing.image.Image(logo)
        sheet.add_image(img, 'G2')


    ### fill the file
    sheet['A2'] = 'Line name:'
    sheet['B2'] = line_name
    sheet['A3'] = 'Date of survey:'

    sheet.freeze_panes = 'C7'
    sheet.sheet_view.zoomScale = 90

    sheet.merge_cells('A5:A6')
    sheet['A5'] = 'From tower'
    sheet.merge_cells('B5:B6')
    sheet['B5'] = 'To tower'
    sheet.merge_cells('C5:C6')
    sheet['C5'] = 'Span length, m'
    sheet.merge_cells('D5:D6')
    sheet['D5'] = 'Conductor Temperature, ºС'
    sheet.merge_cells('E5:E6')
    sheet['E5'] = 'Clearance to infringing tree, m'
    sheet.merge_cells('F5:G5')
    sheet['F5'] = 'Infringing tree location'
    sheet['F6'] = 'Station, m'
    sheet['G6'] = 'Offset, m'
    sheet.merge_cells('H5:H6')
    sheet['H5'] = 'Photography Number'

    for ce in range(len(tab)):
        sheet[str('A' + str(int(7 + ce)))] = tab[ce][8]
        sheet[str('B' + str(int(7 + ce)))] = tab[ce][9]
        sheet[str('C' + str(int(7 + ce)))] = tab[ce][10]
        sheet[str('E' + str(int(7 + ce)))] = float(tab[ce][2])
        sheet[str('F' + str(int(7 + ce)))] = tab[ce][11]
        sheet[str('G' + str(int(7 + ce)))] = tab[ce][12]
        sheet[str('H' + str(int(7 + ce)))] = tab[ce][13]

    ### a little bit of style
    from openpyxl.styles import Border, Side, Alignment, Font

    font1 = Font(name='Arial', size=12)
    font1b = Font(name='Arial', size=12, bold=True)
    font2 = Font(name='Arial', size=10, bold=True)
    font3 = Font(name='Arial', size=10)

    align1 = Alignment(horizontal='center', vertical='center', wrap_text=True)

    bold_left = Border(left=Side(style='medium'),
                       right=Side(style='thin'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))

    bold_right = Border(left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

    bold_mid = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))

    bold_top_mid = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='medium'),
                          bottom=Side(style='thin'))

    bold_bot_mid = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='medium'))

    bold_mid_left = Border(left=Side(style='medium'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    bold_mid_right = Border(left=Side(style='thin'),
                            right=Side(style='medium'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

    bold_left_bot = Border(left=Side(style='medium'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='medium'))

    bold_right_bot = Border(left=Side(style='thin'),
                            right=Side(style='medium'),
                            top=Side(style='thin'),
                            bottom=Side(style='medium'))

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['E'].width = 13
    sheet.column_dimensions['F'].width = 13
    sheet.column_dimensions['G'].width = 13
    sheet.column_dimensions['H'].width = 40
    sheet.row_dimensions[5].height = 30
    # sheet.row_dimensions[6].height = 15

    sheet['A2'].font = font1
    sheet['B2'].font = font1b
    sheet['A3'].font = font1

    sheet['A5'].border = bold_left
    sheet['A6'].border = bold_left
    sheet['B5'].border = bold_mid
    sheet['B6'].border = bold_mid
    sheet['C5'].border = bold_mid
    sheet['C6'].border = bold_mid
    sheet['D5'].border = bold_mid
    sheet['D6'].border = bold_mid
    sheet['E5'].border = bold_mid
    sheet['E6'].border = bold_mid
    sheet['F5'].border = bold_top_mid
    sheet['F6'].border = bold_bot_mid
    sheet['G5'].border = bold_top_mid
    sheet['G6'].border = bold_bot_mid
    sheet['H5'].border = bold_right
    sheet['H6'].border = bold_right

    for row in sheet['A5':'H6']:
        for cell in row:
            cell.alignment = align1
            cell.font = font2

    for row in sheet['A7':str('H' + str(len(tab) + 6))]:
        for cell in row:
            cell.alignment = align1
            cell.font = font3
            cell.border = thin_border

    for col in sheet['C7':str('C' + str(len(tab) + 6))]:
        for cell in col:
            cell.number_format = '0.00'

    for col in sheet['E7':str('G' + str(len(tab) + 6))]:
        for cell in col:
            cell.number_format = '0.00'

    for col in sheet['A7':str('A' + str(len(tab) + 5))]:
        for cell in col:
            cell.border = bold_mid_left

    for col in sheet['H7':str('H' + str(len(tab) + 5))]:
        for cell in col:
            cell.border = bold_mid_right

    for row in sheet[str('B' + str(len(tab) + 6)):str('G' + str(len(tab) + 6))]:
        for cell in row:
            cell.border = bold_bot_mid

    sheet[str('A' + str(len(tab) + 6))].border = bold_left_bot
    sheet[str('H' + str(len(tab) + 6))].border = bold_right_bot

    wb.save(p / str(line_name + '_VCIA_Report_v01.xlsx'))


def write_csv(tabtowrite, path):
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerows(tabtowrite)


def idsforqgis(coords, idsspec):
    # add x, y, z from coords to id_tab and write to file
    a = b = 0
    for a in range(len(idsspec)):
        for b in range(len(coords)):
            if idsspec[a][0] == coords[b][0]:
                for co in [coords[b][1], coords[b][2], coords[b][3]]:
                    idsspec[a].append(co)

    # then looking for dubs (they are with no coords)
    a = b = 0
    for a in range(len(idsspec)):
        if len(idsspec[a]) < 6:
            for b in range(len(idsspec)):
                if idsspec[a][1] == idsspec[b][1]:
                    for co in [idsspec[b][3], idsspec[b][4], idsspec[b][5]]:
                        idsspec[a].append(co)
                    break

    # write tab out
    write_csv(idsspec, (p / 'qgis1.txt'))

    return idsspec


def azimuth(a, b):
    # by two points we get azimuth
    dx = a[0] - b[0]
    dy = a[1] - b[1]
    dist = sqrt(dx*dx + dy*dy)    # dist a to b
    dx2 = abs(dx)
    beta = degrees(acos(dx2/dist))
    if dx > 0:
        if dy < 0:
            angle = 270 + beta
        else:
            angle = 270 - beta
    else:
        if dy < 0:
            angle = 90 - beta
        else:
            angle = 90 + beta

    return round(angle, 2)


def qgis_spans(ngabtab, coords):
    # format and export tab with notgab spans coords
    spantab = []  #   final tab
    templist = {}   # temporary for list of used names
    for a in range(len(ngabtab)):
        spanname = str(f'{ngabtab[a][8]} - {ngabtab[a][9]}')
        if spanname not in templist.keys():
            # check if we already NOT write span
            templist[spanname] = [1, ngabtab[a][2]]
            spantab.append([])
            spantab[-1].append(spanname)
            spantab[-1].append(ngabtab[a][2])   # add 3d dist
            spantab[-1].append(ngabtab[a][10])   # add span length

            az_coo = []  # coords to calc azimuth
            for t in [ngabtab[a][8], ngabtab[a][9]]:
                for b in range(len(coords)):
                    if t == coords[b][1]:
                        for coo in [coords[b][3], coords[b][4], coords[b][5]]:
                            spantab[-1].append(coo)
                            az_coo.append(coo)
            az = azimuth(az_coo[:2], az_coo[3:5])
            spantab[-1].append(az)    # add azimuth
            spantab[-1].append(round((90-az), 2))    # add map angle
        else:
            templist[spanname][0] += 1
            if templist[spanname][1] > ngabtab[a][2]:
                templist[spanname][1] = ngabtab[a][2]

    for sp in range(len(spantab)):
        # from dict write new data to final tab
        if spantab[sp][0] in templist.keys():
            spantab[sp][1] = templist[spantab[sp][0]][1]
            spantab[sp].append(templist[spantab[sp][0]][0])

    write_csv(spantab, (p / 'qgis3.txt'))


# start here
# collect all the files
fz, ctow, spec, logo = filecollect()

# finding line name from name of specification
line_name = spec.stem.split('_')[1]

# start interface
iface()

# read coordinates
s_coo = centerline(ctow)
#print(s_coo[1])

# read specifications and get ids
ids = spec_id(spec)   # get ids
#print(ids[1])

# read files with notgab points and updating the table
ng_tab = notgabread(fz)   # get ngtab (v1)
ng_tab2 = idspannames(ids, ng_tab)    # add id names to ngtab (v2)
ng_tab3 = filltab(ng_tab2, s_coo)    # add calcs (v3)

# printing final tab
# for i in range(len(ng_tab3)):
#     print(ng_tab3[i])

# export to xlsx file
excel_export(ng_tab3)

# start part 2 - creating tables for qgis
print('p a r t _ 2')

ids2 = idsforqgis(s_coo, ids)
print('tab qgis1.txt - done')

write_csv(ng_tab3, (p / 'qgis2.txt'))
print('tab qgis2.txt - done')

qgis_spans(ng_tab3, ids2)
print('tab qgis3.txt - done')


print('\nвсё сработало хорошо :)')
input('нажмите Enter для выхода')


# TODO - добавить обрезку длинных пролетов ?
